# -*- coding: utf-8 -*-
"""オリジナル記録様式 共通ヘルパー（クリーン・メディカルモダン体裁）
   openpyxl で実務に即した様式を生成するための部品。
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

BASE = os.path.dirname(os.path.abspath(__file__))

# ── パレット（サイトと統一：深緑×生成り×金茶）──
FOREST   = "1F4D3A"
FOREST_D = "163A2B"
GOLD     = "A9854A"
GOLD_L   = "EFE7D6"
CREAM    = "F7F4EC"
PAPER    = "FCFAF4"
SAGE     = "4D9B77"
SAGE_L   = "E4EFE7"
INK      = "283129"
INK_MID  = "5C665C"
LINE     = "DAD3C2"
WHITE    = "FFFFFF"
NOTE_BG  = "F6EDDA"

FONT = "Meiryo"   # Excel標準で確実に表示される和文フォント

def _fill(c): return PatternFill("solid", fgColor=c)
def _side(c=LINE, st="thin"): return Side(style=st, color=c)
def _border(c=LINE, st="thin"):
    s=_side(c,st); return Border(left=s,right=s,top=s,bottom=s)

def font(size=10, bold=False, color=INK, italic=False):
    return Font(name=FONT, size=size, bold=bold, color=color, italic=italic)

def align(h="left", v="center", wrap=True, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

def new_sheet(title):
    wb = Workbook(); ws = wb.active; ws.title = title[:31]
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.6, bottom=0.5, header=0.2, footer=0.2)
    return wb, ws

def title_block(ws, text, sub="", last_col="H"):
    """大見出し＋（任意）サブ行。1行目タイトル、2行目サブ。"""
    n = column_index(last_col)
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]; c.value = text
    c.font = font(16, bold=True, color=FOREST_D); c.alignment = align("left","center")
    ws.row_dimensions[1].height = 30
    # 金茶アンダーライン（2行目を細い帯に）
    ws.merge_cells(f"A2:{last_col}2")
    c2 = ws["A2"]; c2.value = sub
    c2.font = font(9, color=INK_MID); c2.alignment = align("left","center")
    ws.row_dimensions[2].height = 16
    for col in range(1, n+1):
        ws.cell(row=2, column=col).border = Border(bottom=_side(GOLD, "medium"))
    return 3  # 次に使える行

def column_index(letter):
    idx=0
    for ch in letter: idx = idx*26 + (ord(ch.upper())-64)
    return idx

def set_widths(ws, widths):
    """widths: {'A':12,'B':30,...}"""
    for col,w in widths.items(): ws.column_dimensions[col].width = w

def label_value_rows(ws, start, pairs, label_col="A", label_w=2, val_w=4, last_col="H", rowh=22):
    """記入欄（ラベル｜記入セル）。pairs=[(label, placeholder or ''),...]
       label_w/val_w はマージ列数。"""
    r = start
    li = column_index(label_col)
    for label, ph in pairs:
        # ラベル
        l_end = get_column_letter(li+label_w-1)
        ws.merge_cells(f"{label_col}{r}:{l_end}{r}")
        c = ws.cell(row=r, column=li); c.value = label
        c.fill = _fill(SAGE_L); c.font = font(10, bold=True, color=FOREST_D)
        c.alignment = align("left","center", indent=1); c.border=_border()
        # 値
        v_start = li+label_w
        v_end = column_index(last_col)
        ws.merge_cells(f"{get_column_letter(v_start)}{r}:{last_col}{r}")
        cv = ws.cell(row=r, column=v_start); cv.value = ph
        cv.font = font(10, color=INK_MID if ph else INK)
        cv.alignment = align("left","center", indent=1); cv.border=_border()
        # マージ範囲の枠線
        for col in range(li, v_end+1):
            ws.cell(row=r, column=col).border = _border()
        ws.row_dimensions[r].height = rowh
        r += 1
    return r

def section(ws, row, text, last_col="H"):
    """セクション帯（深緑背景・白文字）"""
    n=column_index(last_col)
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c=ws.cell(row=row,column=1); c.value=text
    c.fill=_fill(FOREST); c.font=font(11,bold=True,color=WHITE)
    c.alignment=align("left","center",indent=1)
    for col in range(1,n+1): ws.cell(row=row,column=col).fill=_fill(FOREST)
    ws.row_dimensions[row].height=24
    return row+1

def table_header(ws, row, headers, widths=None):
    """表ヘッダー（金茶背景）。headers=[(col_letter,title),...] or [title,...]"""
    for i,h in enumerate(headers):
        col = i+1
        c=ws.cell(row=row,column=col); c.value=h
        c.fill=_fill(GOLD); c.font=font(10,bold=True,color=WHITE)
        c.alignment=align("center","center"); c.border=_border(GOLD,"thin")
    ws.row_dimensions[row].height=22
    return row+1

def table_rows(ws, start, ncols, n=10, rowh=22, sample=None):
    """空の記入行を n 行。sample=最初の数行に薄字の記入例（list of list）"""
    r=start
    for i in range(n):
        for col in range(1,ncols+1):
            c=ws.cell(row=r,column=col); c.border=_border()
            c.alignment=align("left","center",indent=1); c.font=font(10)
        if sample and i < len(sample):
            for col,val in enumerate(sample[i],start=1):
                cc=ws.cell(row=r,column=col); cc.value=val; cc.font=font(9,color=INK_MID,italic=True)
        ws.row_dimensions[r].height=rowh
        r+=1
    return r

def note(ws, row, text, last_col="H", rowh=None):
    """注記ボックス（クリーム背景・金茶左罫）"""
    n=column_index(last_col)
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c=ws.cell(row=row,column=1); c.value=text
    c.fill=_fill(NOTE_BG); c.font=font(9, color=INK_MID)
    c.alignment=align("left","top",wrap=True,indent=1)
    for col in range(1,n+1):
        cell=ws.cell(row=row,column=col); cell.fill=_fill(NOTE_BG)
        cell.border=Border(left=_side(GOLD,"medium"))
    lines = text.count("\n")+1
    ws.row_dimensions[row].height = rowh or max(20, 14*lines+6)
    return row+1

def blank(ws,row,h=8):
    ws.row_dimensions[row].height=h
    return row+1

def save(wb, folder, filename):
    d=os.path.join(BASE, folder); os.makedirs(d, exist_ok=True)
    p=os.path.join(d, filename); wb.save(p)
    print(f"  saved: {folder}/{filename}")
    return p
