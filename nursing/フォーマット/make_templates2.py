"""追加テンプレート生成スクリプト"""
import os, sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from docx import Document
from docx.shared import Pt, Cm, RGBColor

BASE = os.path.dirname(os.path.abspath(__file__))

TEAL="FF0E7490"; WHITE="FFFFFFFF"; LIGHT="FFCFFAFE"; GRAY="FFF8FAFC"

def hf(): return PatternFill("solid", fgColor=TEAL)
def sf(): return PatternFill("solid", fgColor=LIGHT)
def gf(): return PatternFill("solid", fgColor=GRAY)
def hfont(sz=10,col=WHITE): return Font(name="Meiryo UI",size=sz,bold=True,color=col)
def bfont(sz=9,bold=False): return Font(name="Meiryo UI",size=sz,bold=bold)
def bd():
    s=Side(style="thin")
    return Border(left=s,right=s,top=s,bottom=s)
def sw(ws,col,w): ws.column_dimensions[col].width=w
def sr(ws,r,h): ws.row_dimensions[r].height=h

def hrow(ws,row,cols):
    for letter,text,width in cols:
        c=ws[f"{letter}{row}"]
        c.value=text; c.font=hfont(); c.fill=hf()
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border=bd()
        ws.column_dimensions[letter].width=width
    sr(ws,row,30)

def bc(ws,ref,val="",bold=False,ha="left"):
    c=ws[ref]; c.value=val; c.font=bfont(bold=bold)
    c.alignment=Alignment(horizontal=ha,vertical="center",wrap_text=True); c.border=bd()
    return c

def ec(ws,ref,val=""):
    c=ws[ref]; c.value=val
    c.font=Font(name="Meiryo UI",size=9,color="FF6B7280",italic=True)
    c.fill=gf(); c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
    c.border=bd()

def save(wb,folder,fname):
    p=os.path.join(BASE,folder,fname); wb.save(p)
    print(f"OK: {folder}/{fname}")

# ===== 訪問看護記録書I（フェイスシート） =====
def make_facesheet():
    wb=Workbook(); ws=wb.active; ws.title="訪問看護記録書I"

    ws.merge_cells("A1:H1")
    ws["A1"].value="訪問看護記録書Ⅰ（基本情報・フェイスシート）"
    ws["A1"].font=Font(name="Meiryo UI",size=13,bold=True,color=TEAL)
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center"); sr(ws,1,26)

    rows=[
        ("利用者氏名（ふりがな）","（　　　　　　　　）","生年月日","　年　月　日生（　歳）"),
        ("性別","男 ・ 女","要介護認定","なし / 要支援　 / 要介護　"),
        ("障害の種別","身体・精神・知的・難病・その他","障害支援区分","区分　　　/手帳等級："),
        ("住所","〒","電話番号",""),
        ("主傷病名・診断名","","副傷病名",""),
        ("主治医氏名・医療機関","","電話番号",""),
        ("緊急連絡先①（氏名・続柄）","","電話（日中）",""),
        ("緊急連絡先②（氏名・続柄）","","電話（夜間）",""),
        ("担当相談支援専門員","","事業所・電話",""),
        ("ケアマネジャー（介護保険）","","事業所・電話",""),
        ("保険の種類","健保 / 国保 / 後期高齢 / その他","訪問看護指示書有効期限","　　年　月　日まで"),
        ("自立支援医療（精神通院）","あり（上限月額：　　　円） / なし","受給者証番号",""),
        ("アレルギー","","感染症・注意事項",""),
        ("ADL・機能レベル","","認知機能",""),
    ]
    letters="ABCDEFGH"
    for i,(l1,v1,l2,v2) in enumerate(rows,start=2):
        ws.merge_cells(f"A{i}:B{i}"); ws[f"A{i}"].value=l1
        ws.merge_cells(f"C{i}:D{i}"); ws[f"C{i}"].value=v1
        ws.merge_cells(f"E{i}:F{i}"); ws[f"E{i}"].value=l2
        ws.merge_cells(f"G{i}:H{i}"); ws[f"G{i}"].value=v2
        for cell in [f"A{i}",f"E{i}"]:
            ws[cell].font=bfont(bold=True); ws[cell].fill=sf()
            ws[cell].alignment=Alignment(horizontal="right",vertical="center"); ws[cell].border=bd()
        for cell in [f"C{i}",f"G{i}"]:
            ws[cell].font=bfont(); ws[cell].alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
            ws[cell].border=bd()
        sr(ws,i,22)

    # 医療処置・特記事項
    for rn,title,h in [(16,"【医療処置内容（現在実施中のもの）】",60),(18,"【服薬管理・処方薬一覧】",60),(20,"【特記事項・家族への指導事項】",60)]:
        ws.merge_cells(f"A{rn}:H{rn}"); ws[f"A{rn}"].value=title
        ws[f"A{rn}"].font=hfont(sz=9); ws[f"A{rn}"].fill=hf()
        ws[f"A{rn}"].alignment=Alignment(horizontal="left",vertical="center"); ws[f"A{rn}"].border=bd(); sr(ws,rn,20)
        ws.merge_cells(f"A{rn+1}:H{rn+1}"); ws[f"A{rn+1}"].font=bfont()
        ws[f"A{rn+1}"].alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
        ws[f"A{rn+1}"].border=bd(); sr(ws,rn+1,h)

    ws.merge_cells("A22:H22")
    ws["A22"].value="作成日：　　年　月　日　担当看護師：　　　　　　　　最終更新日：　　年　月　日"
    ws["A22"].font=bfont(); ws["A22"].alignment=Alignment(horizontal="left",vertical="center"); ws["A22"].border=bd(); sr(ws,22,24)

    for col,w in zip("ABCDEFGH",[12,12,14,12,12,12,14,12]):
        ws.column_dimensions[col].width=w

    save(wb,"02_訪問看護記録様式","06_訪問看護記録書Ⅰ（フェイスシート）.xlsx")


# ===== 支援記録（日報）A型・B型 =====
def make_nippo():
    wb=Workbook(); ws=wb.active; ws.title="支援記録（日報）"
    ws.page_setup.orientation="landscape"

    ws.merge_cells("A1:K1")
    ws["A1"].value="支援記録（日報）　就労継続支援A型・B型用　　　　年　　月　　日（　曜日）"
    ws["A1"].font=Font(name="Meiryo UI",size=12,bold=True,color=TEAL)
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center"); sr(ws,1,24)

    ws.merge_cells("A2:D2"); ws["A2"].value="天候："
    ws.merge_cells("E2:G2"); ws["E2"].value="開所時間："
    ws.merge_cells("H2:K2"); ws["H2"].value="出勤職員："
    for c in ["A2","E2","H2"]: ws[c].font=bfont(bold=True); ws[c].fill=sf(); ws[c].border=bd(); ws[c].alignment=Alignment(horizontal="right",vertical="center")
    sr(ws,2,20)

    cols=[("A","利用者氏名",12),("B","出欠",6),("C","出勤時刻",8),("D","退勤時刻",8),
          ("E","主な作業内容",22),("F","体調・様子",18),("G","特記事項・申し送り",22),
          ("H","工賃対象\n時間（h）",8),("I","支援者確認",9),("J","管理者確認",9),("K","備考",10)]
    hrow(ws,3,cols)

    ex=["山田 太郎","出","9:00","15:00","封入作業・箱折り（本日担当A）",
        "体調良好。作業集中できていた。午後少し疲れ気味。","家族より電話あり（体調確認）→問題なし","6","鈴木","山田",""]
    for i,v in enumerate(ex): ec(ws,f"{get_column_letter(i+1)}4",str(v))
    sr(ws,4,35)

    for row in range(5,25):
        for col in range(1,12): bc(ws,f"{get_column_letter(col)}{row}")
        sr(ws,row,32)

    # 合計行
    r=25
    ws.merge_cells(f"A{r}:G{r}"); ws[f"A{r}"].value="【本日の特記事項・引き継ぎ事項】"
    ws[f"A{r}"].font=hfont(sz=9); ws[f"A{r}"].fill=hf(); ws[f"A{r}"].alignment=Alignment(horizontal="left",vertical="center"); ws[f"A{r}"].border=bd(); sr(ws,r,20)
    ws.merge_cells(f"A{r+1}:G{r+1}"); ws[f"A{r+1}"].font=bfont(); ws[f"A{r+1}"].alignment=Alignment(horizontal="left",vertical="top",wrap_text=True); ws[f"A{r+1}"].border=bd(); sr(ws,r+1,50)
    ws.merge_cells(f"H{r}:K{r+1}"); ws[f"H{r}"].value="記録者署名：\n\n管理者確認："
    ws[f"H{r}"].font=bfont(); ws[f"H{r}"].alignment=Alignment(horizontal="left",vertical="top",wrap_text=True); ws[f"H{r}"].border=bd()

    save(wb,"03_就労支援記録様式","09_支援記録日報（A型B型）.xlsx")


# ===== シフト表 =====
def make_shift():
    wb=Workbook(); ws=wb.active; ws.title="勤務シフト表"
    ws.page_setup.orientation="landscape"

    ws.merge_cells("A1:AF1")
    ws["A1"].value="勤務シフト表　　　　年　　月分"
    ws["A1"].font=Font(name="Meiryo UI",size=13,bold=True,color=TEAL)
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center"); sr(ws,1,26)

    # 月〜日のヘッダー
    ws["A2"].value="氏名"; ws["A2"].font=hfont(); ws["A2"].fill=hf(); ws["A2"].alignment=Alignment(horizontal="center",vertical="center"); ws["A2"].border=bd(); ws.column_dimensions["A"].width=14
    ws["B2"].value="雇用形態"; ws["B2"].font=hfont(); ws["B2"].fill=hf(); ws["B2"].alignment=Alignment(horizontal="center",vertical="center"); ws["B2"].border=bd(); ws.column_dimensions["B"].width=8

    weekdays=["日","月","火","水","木","金","土"]
    for d in range(1,32):
        col=get_column_letter(d+2)
        c=ws[f"{col}2"]
        c.value=str(d); c.font=hfont(sz=8); c.fill=hf()
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bd()
        ws.column_dimensions[col].width=3.2
    # 合計列
    ws[f"{get_column_letter(33)}2"].value="勤務\n日数"; ws[f"{get_column_letter(33)}2"].font=hfont(sz=8); ws[f"{get_column_letter(33)}2"].fill=hf()
    ws[f"{get_column_letter(33)}2"].alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); ws[f"{get_column_letter(33)}2"].border=bd(); ws.column_dimensions[get_column_letter(33)].width=6
    ws[f"{get_column_letter(34)}2"].value="時間\n合計"; ws[f"{get_column_letter(34)}2"].font=hfont(sz=8); ws[f"{get_column_letter(34)}2"].fill=hf()
    ws[f"{get_column_letter(34)}2"].alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); ws[f"{get_column_letter(34)}2"].border=bd(); ws.column_dimensions[get_column_letter(34)].width=6
    sr(ws,2,30)

    # 記入凡例
    ws.merge_cells("A3:AF3")
    ws["A3"].value="【凡例】 ○=出勤　△=半日　×=公休　有=有給　欠=欠勤　早=早番(6:00〜14:00)　遅=遅番(14:00〜22:00)　夜=夜勤"
    ws["A3"].font=Font(name="Meiryo UI",size=8,color="FF6B7280"); ws["A3"].alignment=Alignment(horizontal="left",vertical="center"); sr(ws,3,18)

    # スタッフ行（15名分）
    ex=[("山田 花子","常勤看護師"),("鈴木 太郎","常勤看護師"),("田中 美穂","非常勤看護師"),("佐藤 健","サビ管")]
    for i,row in enumerate(range(4,19)):
        if i<len(ex): nm,role=ex[i]
        else: nm,role="",""
        ws[f"A{row}"].value=nm; ws[f"A{row}"].font=bfont(bold=i<len(ex)); ws[f"A{row}"].border=bd(); ws[f"A{row}"].alignment=Alignment(horizontal="left",vertical="center")
        ws[f"B{row}"].value=role; ws[f"B{row}"].font=bfont(sz=8); ws[f"B{row}"].border=bd(); ws[f"B{row}"].alignment=Alignment(horizontal="center",vertical="center")
        for d in range(1,35): bc(ws,f"{get_column_letter(d+2)}{row}",ha="center")
        sr(ws,row,20)

    # 常勤換算計算エリア
    cr=20
    ws.merge_cells(f"A{cr}:B{cr}"); ws[f"A{cr}"].value="常勤換算計算"
    ws[f"A{cr}"].font=bfont(bold=True); ws[f"A{cr}"].fill=sf(); ws[f"A{cr}"].alignment=Alignment(horizontal="center",vertical="center"); ws[f"A{cr}"].border=bd(); sr(ws,cr,22)
    ws.merge_cells(f"C{cr}:AF{cr}")
    ws[f"C{cr}"].value="【常勤の基準時間：　　h/週】　常勤換算数 = 各スタッフの週勤務時間 ÷ 常勤基準時間　　合計常勤換算：　　　人（基準：看護師2.5人以上）"
    ws[f"C{cr}"].font=Font(name="Meiryo UI",size=9,color="FFDC2626",bold=True)
    ws[f"C{cr}"].alignment=Alignment(horizontal="left",vertical="center"); ws[f"C{cr}"].border=bd()

    save(wb,"06_事故苦情ヒヤリ様式","07_勤務シフト表.xlsx")


# ===== アセスメント票（就労支援版） =====
def make_assessment():
    wb=Workbook(); ws=wb.active; ws.title="アセスメント票"

    ws.merge_cells("A1:H1")
    ws["A1"].value="アセスメント票（就労継続支援A型・B型・相談支援共通）"
    ws["A1"].font=Font(name="Meiryo UI",size=13,bold=True,color=TEAL)
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center"); sr(ws,1,26)

    ws.merge_cells("A2:H2")
    ws["A2"].value="アセスメント実施日：　　年　月　日　実施者：　　　　　　利用者氏名：　　　　　　（参加者：本人・家族・相談支援専門員・その他：　　　）"
    ws["A2"].font=bfont(); ws["A2"].alignment=Alignment(horizontal="left",vertical="center"); ws["A2"].border=bd(); sr(ws,2,22)

    sections_def=[
        ("【本人の意向・希望】","どんな仕事・活動をしたいか。将来の希望。生活の希望。",70),
        ("【日常生活の状況】","起床・就寝時刻 / 食事・服薬の自己管理 / 清潔管理 / 外出・移動 / コミュニケーション能力",60),
        ("【就労能力・作業能力】","作業の持続時間（　　h） / 集中力 / 体力・持久力 / 手先の器用さ / 指示理解力 / 協調性",60),
        ("【健康・医療の状況】","現在の傷病 / 服薬状況 / 通院状況 / 精神状態の波 / 体調悪化の兆候・サイン",60),
        ("【支援のニーズ（本人・家族）】","何を支援してほしいか。日常生活上の困りごと。就労上の困りごと。",60),
        ("【家族・住環境の状況】","同居家族 / キーパーソン / 家族の支援状況 / 住居の状況",50),
        ("【社会参加・日中活動の状況】","現在の利用サービス / 余暇活動 / 地域との関係",50),
        ("【支援上の課題・優先事項】","アセスメントから見えた課題。個別支援計画の目標設定に向けた優先事項。",70),
    ]
    row=3
    for title,desc,h in sections_def:
        ws.merge_cells(f"A{row}:H{row}"); ws[f"A{row}"].value=title
        ws[f"A{row}"].font=hfont(sz=9); ws[f"A{row}"].fill=hf()
        ws[f"A{row}"].alignment=Alignment(horizontal="left",vertical="center"); ws[f"A{row}"].border=bd(); sr(ws,row,20)
        ws.merge_cells(f"A{row+1}:H{row+1}"); ws[f"A{row+1}"].value=f"（{desc}）"
        ws[f"A{row+1}"].font=Font(name="Meiryo UI",size=9,color="FF6B7280",italic=True)
        ws[f"A{row+1}"].fill=gf(); ws[f"A{row+1}"].alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
        ws[f"A{row+1}"].border=bd(); sr(ws,row+1,h)
        row+=2

    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"].value="アセスメント実施者署名：　　　　　　　　　　　日付：　　年　月　日　確認者（サビ管）：　　　　　　　　"
    ws[f"A{row}"].font=bfont(); ws[f"A{row}"].alignment=Alignment(horizontal="left",vertical="center"); ws[f"A{row}"].border=bd(); sr(ws,row,26)

    for col,w in zip("ABCDEFGH",[14,14,14,14,14,14,14,14]): ws.column_dimensions[col].width=w
    save(wb,"03_就労支援記録様式","10_アセスメント票.xlsx")


# ===== サービス担当者会議記録（相談支援用） =====
def make_kaigi():
    wb=Workbook(); ws=wb.active; ws.title="サービス担当者会議記録"

    ws.merge_cells("A1:G1")
    ws["A1"].value="サービス担当者会議記録（相談支援・計画相談用）"
    ws["A1"].font=Font(name="Meiryo UI",size=13,bold=True,color=TEAL)
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center"); sr(ws,1,26)

    info=[
        ("開催日時","　　年　月　日（　曜日）　　時　　分〜　　時　　分","開催場所",""),
        ("利用者氏名","（　　　）様","進行・記録者",""),
    ]
    for i,(l1,v1,l2,v2) in enumerate(info,start=2):
        ws.merge_cells(f"A{i}:B{i}"); ws[f"A{i}"].value=l1
        ws.merge_cells(f"C{i}:D{i}"); ws[f"C{i}"].value=v1
        ws.merge_cells(f"E{i}:F{i}"); ws[f"E{i}"].value=l2
        ws[f"G{i}"].value=v2
        for cell in [f"A{i}",f"E{i}"]:
            ws[cell].font=bfont(bold=True); ws[cell].fill=sf()
            ws[cell].alignment=Alignment(horizontal="right",vertical="center"); ws[cell].border=bd()
        for cell in [f"C{i}",f"G{i}"]:
            ws[cell].font=bfont(); ws[cell].alignment=Alignment(horizontal="left",vertical="center",wrap_text=True); ws[cell].border=bd()
        sr(ws,i,22)

    # 参加者テーブル
    ws.merge_cells("A4:G4"); ws["A4"].value="【参加者一覧】"
    ws["A4"].font=hfont(sz=9); ws["A4"].fill=hf(); ws["A4"].alignment=Alignment(horizontal="left",vertical="center"); ws["A4"].border=bd(); sr(ws,4,20)

    part_cols=[("A","氏名",16),("B","所属機関",18),("C","役割・職種",14),("D","参加形態",12)]
    hrow(ws,5,part_cols)
    ex_parts=[("山田 太郎","○○相談支援事業所","相談支援専門員（主任）","対面"),
              ("佐藤 花子","○○就労継続支援事業所","サービス管理責任者","対面"),
              ("田中 一郎","○○クリニック","精神科医","欠席（文書参加）"),
              ("利用者本人","—","本人","対面")]
    for i,r in enumerate(ex_parts,start=6):
        for j,v in enumerate(r): ec(ws,f"{get_column_letter(j+1)}{i}",v)
        sr(ws,i,20)
    ws.merge_cells("E5:G5"); ws["E5"].value="備考"; ws["E5"].font=hfont(sz=9); ws["E5"].fill=hf(); ws["E5"].alignment=Alignment(horizontal="center",vertical="center"); ws["E5"].border=bd()
    for r in range(6,10):
        ws.merge_cells(f"E{r}:G{r}"); bc(ws,f"E{r}")

    text_secs=[
        (10,"【検討事項①】個別支援計画・サービス等利用計画の目標の確認",50,"各支援者から利用者の現在の状況・目標に向けた進捗を確認する。"),
        (12,"【検討事項②】サービスの実施状況・課題の共有",60,"各事業所から当月の支援状況・利用者の変化・課題を共有する。"),
        (14,"【検討事項③】計画の変更・調整事項",50,""),
        (16,"【決定事項・次回の目標・役割分担】",60,"次回会議の予定：　　年　月　日（　）　　時　　分〜"),
    ]
    for rn,title,h,hint in text_secs:
        ws.merge_cells(f"A{rn}:G{rn}"); ws[f"A{rn}"].value=title
        ws[f"A{rn}"].font=hfont(sz=9); ws[f"A{rn}"].fill=hf()
        ws[f"A{rn}"].alignment=Alignment(horizontal="left",vertical="center"); ws[f"A{rn}"].border=bd(); sr(ws,rn,20)
        ws.merge_cells(f"A{rn+1}:G{rn+1}"); ws[f"A{rn+1}"].value=hint
        ws[f"A{rn+1}"].font=Font(name="Meiryo UI",size=9,color="FF6B7280",italic=True) if hint else bfont()
        ws[f"A{rn+1}"].fill=gf() if hint else PatternFill()
        ws[f"A{rn+1}"].alignment=Alignment(horizontal="left",vertical="top",wrap_text=True); ws[f"A{rn+1}"].border=bd(); sr(ws,rn+1,h)

    for col,w in zip("ABCDEFG",[16,18,14,12,12,12,14]): ws.column_dimensions[col].width=w
    save(wb,"04_相談支援様式","03_サービス担当者会議記録.xlsx")


# ===== 利用者負担額管理表 =====
def make_futankanri():
    wb=Workbook(); ws=wb.active; ws.title="利用者負担管理"

    ws.merge_cells("A1:H1")
    ws["A1"].value="利用者負担額管理表　　　　年　　月分"
    ws["A1"].font=Font(name="Meiryo UI",size=12,bold=True,color=TEAL)
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center"); sr(ws,1,24)

    ws.merge_cells("A2:H2")
    ws["A2"].value="※ 月額上限管理が必要な方は「上限管理表」（国保連様式）を別途作成してください。"
    ws["A2"].font=Font(name="Meiryo UI",size=8,color="FFDC2626",bold=True)
    ws["A2"].alignment=Alignment(horizontal="left",vertical="center"); sr(ws,2,18)

    cols=[("A","利用者氏名",14),("B","保険種別",10),("C","負担割合",8),
          ("D","月額上限額（円）",12),("E","サービス費用合計（円）",14),
          ("F","利用者負担額（円）",14),("G","領収日",10),("H","備考",14)]
    hrow(ws,3,cols)

    ex=[["山田 太郎","健康保険","1割（自立支援）",5000,32000,3200,"2026/5/25","精神通院医療適用"],
        ["佐藤 花子","国民健康保険","1割",37200,37200,3720,"2026/5/25",""]]
    for i,row in enumerate(ex,start=4):
        for j,v in enumerate(row): ec(ws,f"{get_column_letter(j+1)}{i}",str(v))
        sr(ws,i,22)

    for row in range(6,26):
        for col in range(1,9): bc(ws,f"{get_column_letter(col)}{row}",ha="right" if col in [3,4,5,6] else "left")
        sr(ws,row,22)

    save(wb,"06_事故苦情ヒヤリ様式","08_利用者負担額管理表.xlsx")


# ===== 体制届チェックシート =====
def make_taiseicheck():
    wb=Workbook(); ws=wb.active; ws.title="体制届チェックシート"

    ws.merge_cells("A1:F1")
    ws["A1"].value="体制届チェックシート（加算算定管理用）"
    ws["A1"].font=Font(name="Meiryo UI",size=13,bold=True,color=TEAL)
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center"); sr(ws,1,26)

    ws.merge_cells("A2:F2")
    ws["A2"].value="【重要】体制届を出さずに加算を算定すると不正請求となります。算定開始月の前月末日までに届出が必要です。"
    ws["A2"].font=Font(name="Meiryo UI",size=9,color="FFDC2626",bold=True)
    ws["A2"].alignment=Alignment(horizontal="left",vertical="center"); sr(ws,2,20)

    cols=[("A","加算名",28),("B","単位数",10),("C","届出提出日",12),
          ("D","算定開始月",10),("E","要件充足確認",16),("F","次回更新・注意事項",20)]
    hrow(ws,3,cols)

    items=[
        ["処遇改善加算（Ⅰ）","基本報酬×約15-20%","毎年4/15","毎年6月〜","計画書提出・職場環境整備確認","毎年更新（7/31実績報告）"],
        ["福祉専門職員配置等加算（Ⅰ）","15単位/日","","","有資格者比率25%以上を確認","職員変更時に再確認"],
        ["24時間対応体制加算","月6,400円","","","24時間連絡・緊急訪問体制整備","体制変更時に届出変更必要"],
        ["目標工賃達成加算（B型）","10単位/日","","","前年度平均工賃が目標値以上","毎年度確認"],
        ["強度行動障害支援者配置加算","7〜42単位/日","","","研修修了者の配置確認","職員変更時に再確認"],
        ["精神障害者支援体制加算","7〜28単位/日","","","研修修了者の配置確認","職員変更時に再確認"],
    ]
    for i,row in enumerate(items,start=4):
        for j,v in enumerate(row): bc(ws,f"{get_column_letter(j+1)}{i}",v)
        sr(ws,i,28)

    for col,w in zip("ABCDEF",[28,10,12,10,16,20]): ws.column_dimensions[col].width=w
    save(wb,"06_事故苦情ヒヤリ様式","09_体制届チェックシート.xlsx")


if __name__=="__main__":
    print("=== 追加テンプレート生成 ===")
    print("\n--- 02_訪問看護記録様式 ---")
    make_facesheet()
    print("\n--- 03_就労支援記録様式 ---")
    make_nippo()
    make_assessment()
    print("\n--- 04_相談支援様式 ---")
    make_kaigi()
    print("\n--- 06_事故苦情ヒヤリ様式 ---")
    make_shift()
    make_futankanri()
    make_taiseicheck()
    print("\n=== 完了 ===")
