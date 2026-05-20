"""
beqd1106 福祉ガイド — Python API サーバー (Render.com)
Excel書類生成 + 将来：Google Sheets連携・グラフ生成
"""

from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment
)
from openpyxl.utils import get_column_letter
import io, re
from datetime import datetime
from typing import Optional

# ================================================================
app = FastAPI(title="福祉ガイド API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ================================================================
# リクエストモデル
# ================================================================
class ExcelRequest(BaseModel):
    feature: str
    raw_text: str
    form_data: dict = {}

# ================================================================
# カラー定数・スタイルヘルパー
# ================================================================
C_CORAL_DARK  = "D4784C"
C_CORAL       = "F0956A"
C_CORAL_LIGHT = "FDE8D8"
C_CREAM       = "FEF6E0"
C_LAVENDER    = "EDE8F8"
C_LAVENDER_DK = "6040A0"
C_MINT        = "E0F2E8"
C_MINT_DK     = "2E7D32"
C_SKY         = "DFF0FB"
C_SKY_DK      = "005A8E"
C_WHITE       = "FFFFFF"
C_DARK        = "2A2020"
C_MID         = "7A6560"

def fill(color):
    return PatternFill("solid", fgColor=color)

def border(color="D0C0B0", style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def left_accent(accent=C_CORAL, inner="D0C0B0"):
    return Border(
        left=Side(style="medium", color=accent),
        right=Side(style="thin",  color=inner),
        top=Side(style="thin",    color=inner),
        bottom=Side(style="thin", color=inner),
    )

def font(size=10, bold=False, color=C_DARK, name="Meiryo"):
    return Font(name=name, size=size, bold=bold, color=color)

def align(h="left", v="top", wrap=True, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

# ================================================================
# Markdownパーサー
# ================================================================
def parse_sections(text: str):
    sections, cur = [], None
    for line in text.split("\n"):
        m2 = re.match(r"^##\s+(.+)", line)
        m3 = re.match(r"^###\s+(.+)", line)
        if m2:
            if cur: sections.append(cur)
            cur = {"title": m2.group(1).strip(), "lines": [], "level": 2}
        elif m3:
            if cur: sections.append(cur)
            cur = {"title": m3.group(1).strip(), "lines": [], "level": 3}
        elif cur:
            ln = re.sub(r"\*\*(.+?)\*\*", r"\1", line).strip()
            if ln:
                cur["lines"].append(ln)
    if cur:
        sections.append(cur)
    return sections

# ================================================================
# 共通：ヘッダー行
# ================================================================
def write_header(ws, title: str, subtitle: str, date_str: str, num_cols: int = 6):
    # タイトル行
    ws.merge_cells(f"A1:{get_column_letter(num_cols)}1")
    c = ws["A1"]
    c.value = title
    c.font = Font(name="Meiryo", size=14, bold=True, color=C_WHITE)
    c.fill = fill(C_CORAL_DARK)
    c.alignment = align(h="center", v="center")
    ws.row_dimensions[1].height = 36

    # サブタイトル行
    ws.merge_cells(f"A2:{get_column_letter(num_cols)}2")
    c = ws["A2"]
    c.value = subtitle
    c.font = font(size=9, color=C_MID)
    c.fill = fill(C_CORAL_LIGHT)
    c.alignment = align(h="center", v="center")
    ws.row_dimensions[2].height = 16

    # メタ情報行
    meta = [("作成日", date_str, "事業所名", "", "担当者", "")]
    for i, row_vals in enumerate(meta, 3):
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=i, column=ci, value=val)
            is_label = (ci % 2 == 1)
            c.font   = font(size=9, bold=is_label)
            c.fill   = fill(C_CORAL_LIGHT if is_label else C_WHITE)
            c.border = border()
            c.alignment = align(v="center")
        ws.row_dimensions[i].height = 18

    # 注意文
    ws.merge_cells(f"A4:{get_column_letter(num_cols)}4")
    c = ws["A4"]
    c.value = "※ AI（Llama 3.1）が生成した下書きです。内容は必ず担当者が確認・修正してください。"
    c.font  = Font(name="Meiryo", size=8, color="C0504F", italic=True)
    c.fill  = fill("FDE8E8")
    c.alignment = align(v="center", indent=1)
    ws.row_dimensions[4].height = 14

    # 空白行
    ws.row_dimensions[5].height = 8
    return 6  # 次の行番号

# ================================================================
# 共通：セクション書き込み
# ================================================================
def write_section(ws, title: str, lines: list, row: int, num_cols: int = 6,
                  section_color=C_CORAL_LIGHT, accent=C_CORAL_DARK) -> int:
    # セクション見出し
    ws.merge_cells(f"A{row}:{get_column_letter(num_cols)}{row}")
    c = ws[f"A{row}"]
    c.value = f"■  {title}"
    c.font  = Font(name="Meiryo", size=10, bold=True, color=accent)
    c.fill  = fill(section_color)
    c.border = left_accent(accent=accent)
    c.alignment = align(v="center", indent=1)
    ws.row_dimensions[row].height = 22
    row += 1

    # コンテンツ
    content = "\n".join(lines)
    if content:
        ws.merge_cells(f"A{row}:{get_column_letter(num_cols)}{row}")
        c = ws[f"A{row}"]
        c.value = content
        c.font  = font(size=10)
        c.fill  = fill(C_WHITE)
        c.border = border()
        c.alignment = align(wrap=True)
        # 行の高さを内容量に応じて調整（最大250px）
        line_count = max(3, content.count("\n") + len(content) // 55 + 1)
        ws.row_dimensions[row].height = min(15 * line_count, 250)
        row += 1

    # 区切り空白
    ws.row_dimensions[row].height = 5
    return row + 1

# ================================================================
# 共通：入力フィールド行（2列ペア）
# ================================================================
def write_field_row(ws, fields: list, row: int, num_cols: int = 6) -> int:
    """fields = [('ラベル', '値'), ('ラベル2', '値2'), ...]"""
    cols_per_field = num_cols // len(fields)
    col = 1
    for label, value in fields:
        lc = ws.cell(row=row, column=col, value=label)
        lc.font  = font(size=9, bold=True)
        lc.fill  = fill(C_CORAL_LIGHT)
        lc.border = border()
        lc.alignment = align(v="center", indent=1)
        col += 1

        # 値セルを残り列に結合
        val_end = col + cols_per_field - 2
        if val_end >= col:
            ws.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(val_end)}{row}")
        vc = ws.cell(row=row, column=col, value=value)
        vc.font  = font(size=10)
        vc.fill  = fill(C_WHITE)
        vc.border = border()
        vc.alignment = align(v="center", wrap=True)
        col = val_end + 1

    ws.row_dimensions[row].height = 18
    return row + 1

# ================================================================
# 書類タイプ別 生成関数
# ================================================================

def gen_plan_shuro(wb, raw_text, fd, date_str):
    ws = wb.active
    ws.title = "個別支援計画書"
    service = fd.get("serviceType", "就労継続支援B型")
    for i, w in enumerate([1, 22, 1, 22, 1, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = write_header(ws, f"個別支援計画書（{service}）",
                       "障害者総合支援法 基準省令第58条 準拠", date_str)
    # 利用者情報
    row = write_field_row(ws, [("利用者氏名", ""), ("年齢・性別", fd.get("age",""))], row)
    row = write_field_row(ws, [("障害種別・診断", fd.get("diagnosis","")), ("障害支援区分", fd.get("level",""))], row)
    row = write_field_row(ws, [("計画期間", fd.get("period","")), ("計画作成者", "")], row)
    ws.row_dimensions[row].height = 8; row += 1

    # AI出力セクション
    for s in parse_sections(raw_text):
        row = write_section(ws, s["title"], s["lines"], row)

    # 同意欄
    ws.row_dimensions[row].height = 8; row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].value = "本計画書の内容について説明を受け、同意します。"
    ws[f"A{row}"].font  = font(size=9, bold=True)
    ws[f"A{row}"].fill  = fill(C_CREAM)
    ws.row_dimensions[row].height = 16; row += 1
    for label, col in [("利用者 署名・捺印", 1), ("保護者 署名・捺印", 3), ("サビ管 署名・捺印", 5)]:
        c = ws.cell(row=row, column=col, value=label)
        c.font = font(size=9, bold=True); c.fill = fill(C_CORAL_LIGHT); c.border = border()
        c.alignment = align(v="center", indent=1)
        end_col = col + 1
        ws.merge_cells(f"{get_column_letter(col+1)}{row}:{get_column_letter(end_col)}{row}")
        ws.cell(row=row, column=col+1).border = border()
    ws.row_dimensions[row].height = 40


def gen_plan_kyotaku(wb, raw_text, fd, date_str):
    ws = wb.active
    ws.title = "個別支援計画書（居宅）"
    for i, w in enumerate([1, 22, 1, 22, 1, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    services = fd.get("services", [])
    svc_str = "・".join(services) if services else "居宅介護"
    row = write_header(ws, f"個別支援計画書（{svc_str}）",
                       "基準省令第24条 準拠", date_str)
    row = write_field_row(ws, [("利用者氏名", ""), ("年齢・性別", fd.get("age",""))], row)
    row = write_field_row(ws, [("障害支援区分", fd.get("level","")), ("障害の状況", fd.get("diagnosis",""))], row)
    row = write_field_row(ws, [("サービス提供時間", fd.get("schedule","")), ("担当者", "")], row)
    ws.row_dimensions[row].height = 8; row += 1

    for s in parse_sections(raw_text):
        row = write_section(ws, s["title"], s["lines"], row)

    # 同意欄
    ws.row_dimensions[row].height = 8; row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].value = "本計画書の内容について説明を受け、同意します。（法令必須）"
    ws[f"A{row}"].font  = font(size=9, bold=True, color="C0504F")
    ws[f"A{row}"].fill  = fill("FDE8E8")
    ws.row_dimensions[row].height = 16; row += 1
    for label, col in [("利用者 署名", 1), ("保護者 署名", 3), ("サ責 確認", 5)]:
        c = ws.cell(row=row, column=col, value=label)
        c.font = font(size=9, bold=True); c.fill = fill(C_CORAL_LIGHT); c.border = border()
        ws.cell(row=row, column=col+1).border = border()
    ws.row_dimensions[row].height = 40


def gen_assessment(wb, raw_text, fd, date_str):
    ws = wb.active
    ws.title = "アセスメントシート"
    for i, w in enumerate([20, 30, 30, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = write_header(ws, "アセスメントシート（ICF準拠）",
                       "ICF：国際生活機能分類に基づくストレングスベースアセスメント", date_str, num_cols=4)
    row = write_field_row(ws, [("利用者氏名", ""), ("利用サービス", fd.get("service",""))], row, num_cols=4)
    row = write_field_row(ws, [("基本情報", fd.get("basic","")), ("障害支援区分", "")], row, num_cols=4)
    ws.row_dimensions[row].height = 8; row += 1

    # ICF 4列テーブルヘッダー
    headers = ["評価領域", "現状・できること", "課題・ニーズ", "強み・リソース"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font  = font(size=10, bold=True, color=C_WHITE)
        c.fill  = fill(C_CORAL_DARK)
        c.border = border(C_CORAL_DARK, "medium")
        c.alignment = align(h="center", v="center")
    ws.row_dimensions[row].height = 22; row += 1

    # 各セクションを行に
    sections = parse_sections(raw_text)
    for i, s in enumerate(sections):
        bg = C_WHITE if i % 2 == 0 else "FFF8F5"
        content = "\n".join(s["lines"])
        # 現状/課題/強みを簡易分割
        genjo, kadai, tsuyomi = "", "", ""
        mode = 0
        for l in s["lines"]:
            if re.search(r"^現状|^状況", l) and len(l) < 20: mode = 1; continue
            if re.search(r"^課題|^ニーズ", l) and len(l) < 20: mode = 2; continue
            if re.search(r"^強み|^リソース", l) and len(l) < 20: mode = 3; continue
            if   mode == 1: genjo   += l + "\n"
            elif mode == 2: kadai   += l + "\n"
            elif mode == 3: tsuyomi += l + "\n"
            else:           genjo   += l + "\n"

        cells_vals = [s["title"], genjo.strip() or content, kadai.strip(), tsuyomi.strip()]
        for ci, val in enumerate(cells_vals, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.font  = font(size=10, bold=(ci == 1))
            c.fill  = fill(C_CORAL_LIGHT if ci == 1 else bg)
            c.border = border()
            c.alignment = align(wrap=True)
        line_count = max(3, content.count("\n") + len(content) // 40 + 1)
        ws.row_dimensions[row].height = min(15 * line_count, 200)
        row += 1


def gen_monitoring(wb, raw_text, fd, date_str):
    ws = wb.active
    ws.title = "モニタリング報告書"
    for i, w in enumerate([1, 25, 1, 25, 1, 25], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = write_header(ws, "モニタリング報告書",
                       "個別支援計画 実施状況の確認", date_str)
    row = write_field_row(ws, [("サービス種別", fd.get("service","")), ("モニタリング期間", fd.get("period",""))], row)
    row = write_field_row(ws, [("実施日", ""), ("参加者", "本人・担当者")], row)
    ws.row_dimensions[row].height = 8; row += 1

    for s in parse_sections(raw_text):
        row = write_section(ws, s["title"], s["lines"], row)

    # 総合評価欄
    ws.row_dimensions[row].height = 8; row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].value = "■  総合評価・計画変更の要否"
    ws[f"A{row}"].font  = font(size=10, bold=True, color=C_CORAL_DARK)
    ws[f"A{row}"].fill  = fill(C_CORAL_LIGHT)
    ws.row_dimensions[row].height = 22; row += 1
    ws.merge_cells(f"A{row}:F{row+2}")
    ws[f"A{row}"].border = border()
    ws[f"A{row}"].alignment = align()
    ws.row_dimensions[row].height = 60


def gen_nursing_report(wb, raw_text, fd, date_str):
    ws = wb.active
    ws.title = "看護記録"
    for i, w in enumerate([18, 20, 18, 20, 12, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    record_type = fd.get("recordType", "看護記録・看護サマリー")
    row = write_header(ws, record_type, "訪問看護 記録・報告書類", date_str)
    row = write_field_row(ws, [("対象者", fd.get("patient","")), ("主治医", "")], row)

    # バイタルサイン表
    ws.row_dimensions[row].height = 8; row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].value = "■  バイタルサイン"
    ws[f"A{row}"].font  = font(size=10, bold=True, color=C_SKY_DK)
    ws[f"A{row}"].fill  = fill(C_SKY)
    ws[f"A{row}"].border = left_accent(accent=C_SKY_DK)
    ws.row_dimensions[row].height = 22; row += 1

    vitals = [
        ("血圧 (mmHg)", fd.get("bp",""), "脈拍 (回/分)", fd.get("pulse",""),   "体温 (℃)",   fd.get("temp","")),
        ("SpO2 (%)",    fd.get("spo2",""), "意識レベル",  "",                    "呼吸数",      ""),
    ]
    for vrow in vitals:
        for ci, val in enumerate(vrow, 1):
            c = ws.cell(row=row, column=ci, value=val)
            is_label = (ci % 2 == 1)
            c.font   = font(size=9, bold=is_label)
            c.fill   = fill(C_SKY if is_label else C_WHITE)
            c.border = border()
            c.alignment = align(h="center" if is_label else "left", v="center")
        ws.row_dimensions[row].height = 18; row += 1

    ws.row_dimensions[row].height = 8; row += 1
    for s in parse_sections(raw_text):
        row = write_section(ws, s["title"], s["lines"], row,
                            section_color=C_SKY, accent=C_SKY_DK)


def gen_check_report(wb, raw_text, fd, date_str, title, feature):
    ws = wb.active
    ws.title = title[:31]
    for i, w in enumerate([6, 32, 62], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = write_header(ws, title, "算定要件・請求内容の事前確認", date_str, num_cols=3)
    row = write_field_row(ws, [("事業所名", ""), ("診断日", date_str)], row, num_cols=3)
    ws.row_dimensions[row].height = 8; row += 1

    # 判定表ヘッダー
    for ci, (h, w_) in enumerate([("判定", 6), ("項目", 32), ("内容・根拠・対応")], 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font  = font(10, True, C_WHITE)
        c.fill  = fill(C_CORAL_DARK)
        c.border = border(C_CORAL_DARK, "medium")
        c.alignment = align(h="center", v="center")
    ws.row_dimensions[row].height = 22; row += 1

    MARKS = {
        "✅": C_MINT, "⚠": "FFF3CD", "🔴": "FDE8E8",
        "🎯": C_SKY,  "❌": "FDE8E8", "📊": C_LAVENDER, "🔧": C_CREAM,
    }
    for s in parse_sections(raw_text):
        mark = ""
        bg   = C_WHITE
        for m, c_ in MARKS.items():
            if m in s["title"]:
                mark = m; bg = c_; break
        clean_title = re.sub(r"[^ -~　-鿿＀-￯\s]", "", s["title"]).strip()
        content = "\n".join(s["lines"])

        ws.cell(row=row, column=1, value=mark).alignment = align(h="center")
        ws.cell(row=row, column=2, value=clean_title).font = font(10, True)
        ws.cell(row=row, column=3, value=content).alignment = align(wrap=True)
        for ci in range(1, 4):
            ws.cell(row=row, column=ci).fill   = fill(bg)
            ws.cell(row=row, column=ci).border = border()
        ws.row_dimensions[row].height = max(22, min(15 * (content.count("\n") + 2), 150))
        row += 1

    ws.row_dimensions[row].height = 6; row += 1


def gen_incident(wb, raw_text, fd, date_str):
    ws = wb.active
    ws.title = "事故報告書"
    for i, w in enumerate([1, 22, 1, 22, 1, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = write_header(ws, "事故・ヒヤリハット報告書",
                       "指定障害福祉サービス事業所 行政報告用", date_str)
    row = write_field_row(ws, [("事故種別", fd.get("incidentType","")), ("重傷度", fd.get("level",""))], row)
    row = write_field_row(ws, [("発生日時・場所", fd.get("when","")), ("報告者", "")], row)
    row = write_field_row(ws, [("当事者の状況", fd.get("person","")), ("スタッフ状況", fd.get("staffStatus",""))], row)
    ws.row_dimensions[row].height = 8; row += 1

    for s in parse_sections(raw_text):
        row = write_section(ws, s["title"], s["lines"], row,
                            section_color="FDE8E8", accent="B71C1C")

    # 確認サイン欄
    ws.row_dimensions[row].height = 8; row += 1
    for label, col in [("管理者 確認", 1), ("報告者 署名", 3), ("行政報告日", 5)]:
        c = ws.cell(row=row, column=col, value=label)
        c.font = font(9, True); c.fill = fill(C_CORAL_LIGHT); c.border = border()
        ws.cell(row=row, column=col+1).border = border()
    ws.row_dimensions[row].height = 35


def gen_wage_plan(wb, raw_text, fd, date_str):
    ws = wb.active
    ws.title = "工賃向上計画書"
    for i, w in enumerate([1, 25, 1, 25, 1, 25], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = write_header(ws, "工賃向上計画書（就労継続支援B型）",
                       "障害者総合支援法 / 目標工賃達成加算 算定要件書類", date_str)
    row = write_field_row(ws, [("提出先", "京都府 障害者支援課"), ("提出期限", "毎年4月15日")], row)

    # 目標工賃テーブル
    ws.row_dimensions[row].height = 8; row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].value = "■  工賃目標値"
    ws[f"A{row}"].font  = font(10, True, C_MINT_DK)
    ws[f"A{row}"].fill  = fill(C_MINT)
    ws[f"A{row}"].border = left_accent(accent=C_MINT_DK)
    ws.row_dimensions[row].height = 22; row += 1

    wage_data = [
        ("現在の平均月額工賃", fd.get("wageNow","") + " 円", "目標月額工賃", fd.get("wageTarget","") + " 円"),
        ("登録利用者数", fd.get("users",""), "計画年度", fd.get("year","")),
    ]
    for r_data in wage_data:
        row = write_field_row(ws, [(r_data[0], r_data[1]), (r_data[2], r_data[3])], row)

    ws.row_dimensions[row].height = 8; row += 1
    for s in parse_sections(raw_text):
        row = write_section(ws, s["title"], s["lines"], row,
                            section_color=C_MINT, accent=C_MINT_DK)


def gen_staff_sim(wb, raw_text, fd, date_str):
    ws = wb.active
    ws.title = "人員配置診断"
    for i, w in enumerate([22, 18, 14, 14, 1, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = write_header(ws, "人員配置シミュレーション診断レポート",
                       "2024年度 法定人員基準 充足状況確認", date_str, num_cols=6)
    row = write_field_row(ws, [("登録利用者数", fd.get("users","") + " 名"), ("日平均利用", fd.get("daily","") + " 名"), ("定員", fd.get("capacity","") + " 名")], row)
    ws.row_dimensions[row].height = 8; row += 1

    # 常勤換算計算シート
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"].value = "■  常勤換算計算シート（常勤所定労働時間：週＿＿時間）"
    ws[f"A{row}"].font  = font(10, True, C_LAVENDER_DK)
    ws[f"A{row}"].fill  = fill(C_LAVENDER)
    ws[f"A{row}"].border = left_accent(accent=C_LAVENDER_DK)
    ws.row_dimensions[row].height = 22; row += 1

    sheet_headers = ["職種・氏名", "資格・経験年数", "週勤務時間", "常勤換算値", "", "備考"]
    for ci, h in enumerate(sheet_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font  = font(9, True, C_WHITE)
        c.fill  = fill(C_LAVENDER_DK)
        c.border = border()
        c.alignment = align(h="center", v="center")
    ws.row_dimensions[row].height = 20; row += 1

    for _ in range(6):
        for ci in range(1, 7):
            ws.cell(row=row, column=ci).border = border()
        ws.row_dimensions[row].height = 18; row += 1

    # 合計行
    ws.cell(row=row, column=1, value="合計").font = font(10, True)
    ws.cell(row=row, column=1).fill = fill(C_LAVENDER)
    for ci in range(1, 7):
        ws.cell(row=row, column=ci).border = border()
    ws.row_dimensions[row].height = 20; row += 1

    ws.row_dimensions[row].height = 8; row += 1
    for s in parse_sections(raw_text):
        row = write_section(ws, s["title"], s["lines"], row,
                            section_color=C_LAVENDER, accent=C_LAVENDER_DK)


def gen_easy_read(wb, raw_text, fd, date_str):
    ws = wb.active
    ws.title = "わかりやすい説明文"
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 80

    row = write_header(ws, "わかりやすい説明文", fd.get("purpose",""), date_str, num_cols=2)
    row = write_field_row(ws, [("対象読者", fd.get("reader","")), ("目的", fd.get("purpose",""))], row, num_cols=2)
    ws.row_dimensions[row].height = 8; row += 1

    for s in parse_sections(raw_text):
        row = write_section(ws, s["title"], s["lines"], row, num_cols=2)


def gen_generic(wb, raw_text, title, fd, date_str):
    ws = wb.active
    ws.title = title[:31]
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70

    row = write_header(ws, title, "", date_str, num_cols=2)
    for s in parse_sections(raw_text):
        row = write_section(ws, s["title"], s["lines"], row, num_cols=2)


# ================================================================
# 書類タイトルマッピング
# ================================================================
TITLES = {
    "plan_shuro":     "個別支援計画書（就労継続支援）",
    "plan_kyotaku":   "個別支援計画書（居宅介護）",
    "assessment":     "アセスメントシート",
    "monitoring":     "モニタリング報告書",
    "nursing_report": "看護記録・看護サマリー",
    "kasan_check":    "加算算定診断レポート",
    "claim_check":    "請求エラーチェックレポート",
    "law_summary":    "法令Q&A回答",
    "incident_docs":  "事故報告書",
    "wage_plan":      "工賃向上計画書",
    "easy_read":      "わかりやすい説明文",
    "staff_sim":      "人員配置診断レポート",
}

# ================================================================
# メインエンドポイント
# ================================================================
@app.post("/fill-excel")
async def fill_excel(req: ExcelRequest):
    today     = datetime.now()
    date_str  = f"{today.year}年{today.month}月{today.day}日"
    file_date = today.strftime("%Y%m%d")
    title     = TITLES.get(req.feature, "AI生成書類")

    wb = Workbook()

    if   req.feature == "plan_shuro":     gen_plan_shuro(wb, req.raw_text, req.form_data, date_str)
    elif req.feature == "plan_kyotaku":   gen_plan_kyotaku(wb, req.raw_text, req.form_data, date_str)
    elif req.feature == "assessment":     gen_assessment(wb, req.raw_text, req.form_data, date_str)
    elif req.feature == "monitoring":     gen_monitoring(wb, req.raw_text, req.form_data, date_str)
    elif req.feature == "nursing_report": gen_nursing_report(wb, req.raw_text, req.form_data, date_str)
    elif req.feature in ("kasan_check", "claim_check"):
        gen_check_report(wb, req.raw_text, req.form_data, date_str, title, req.feature)
    elif req.feature == "incident_docs":  gen_incident(wb, req.raw_text, req.form_data, date_str)
    elif req.feature == "wage_plan":      gen_wage_plan(wb, req.raw_text, req.form_data, date_str)
    elif req.feature == "staff_sim":      gen_staff_sim(wb, req.raw_text, req.form_data, date_str)
    elif req.feature == "easy_read":      gen_easy_read(wb, req.raw_text, req.form_data, date_str)
    else:
        gen_generic(wb, req.raw_text, title, req.form_data, date_str)

    # フッター行（全シート共通）
    ws = wb.active
    last_row = ws.max_row + 2
    ws.merge_cells(f"A{last_row}:{get_column_letter(ws.max_column)}{last_row}")
    c = ws[f"A{last_row}"]
    c.value = f"出力元：障碍者福祉事業所 運営ガイド | beqd1106.com/nursing/ | {date_str}"
    c.font  = Font(name="Meiryo", size=8, color="AAAAAA", italic=True)
    c.alignment = Alignment(horizontal="right")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_title = title.replace("/", "／").replace("\\", "")
    filename   = f"{safe_title}_{file_date}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}",
            "Access-Control-Allow-Origin": "*",
        },
    )


@app.get("/health")
def health():
    return {"status": "ok", "service": "beqd1106-fukushi-api"}
